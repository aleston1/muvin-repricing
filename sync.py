"""Sincronización multi-plataforma Muvin.

Compara el stock del ERP (Hansa, espejado en Google Sheets) contra las
publicaciones de Mercado Libre y Tiendanube por SKU raíz, y permite
publicar los productos faltantes con descripción estilo Muvin y fotos
aprobadas manualmente.
"""
from flask import Blueprint, jsonify, request
import requests
import base64
import csv
import io
import json
import os
import re
import urllib.parse

sync_bp = Blueprint("sync", __name__, url_prefix="/api/sync")

ML_BASE   = "https://api.mercadolibre.com"
TN_BASE   = "https://api.tiendanube.com/v1"
TN_UA     = os.environ.get("TN_USER_AGENT", "MuvinSync (aleston@muvin.com.ar)")

SHEET_ID  = os.environ.get("STOCK_SHEET_ID", "1ReMzkvfBbPYNmIcJLGQaCivvsf2-IZ2CadtRlcUgkZE")
SHEET_GID = os.environ.get("STOCK_SHEET_GID", "1134850759")

# Planilla "Productos pendientes de publicar": equivalencias de variantes,
# categorías de Tiendanube, marcas y URLs del fabricante.
EQUIV_SHEET_ID   = os.environ.get("EQUIV_SHEET_ID", "1eFOSU_uXME4AzqZs_-hJkB16xtEY82qiPO5KO2uCXRo")
EQUIV_CACHE_PATH = os.path.join(os.path.dirname(__file__), "equiv_cache.json")

# Cache en disco del último stock parseado (Heroku lo pierde al reiniciar,
# igual que costos.json — la fuente de verdad sigue siendo la planilla).
STOCK_CACHE_PATH = os.path.join(os.path.dirname(__file__), "stock_cache.json")


# ---------------------------------------------------------------- helpers ML

def ml_get(path, token, params=None):
    r = requests.get(ML_BASE + path, headers={"Authorization": f"Bearer {token}"},
                     params=params, timeout=20)
    r.raise_for_status()
    return r.json()


def ml_post(path, token, body):
    r = requests.post(ML_BASE + path, headers={"Authorization": f"Bearer {token}"},
                      json=body, timeout=30)
    return r


# ------------------------------------------------------------ fotos (re-hosting)
# Muchos sitios bloquean que terceros descarguen sus imágenes (hotlink).
# Bajamos la imagen nosotros con UA de navegador y la subimos como archivo
# propio a cada plataforma.

BROWSER_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
              "(KHTML, like Gecko) Chrome/125.0 Safari/537.36")


def descargar_imagen(url):
    # Imágenes subidas por el usuario desde su compu: llegan como data URI
    # (data:image/jpeg;base64,....). Se decodifican sin ir a la red.
    if isinstance(url, str) and url.startswith("data:"):
        cabecera, _, datos = url.partition(",")
        ct = "image/jpeg"
        m = re.match(r"data:([^;]+)", cabecera)
        if m:
            ct = m.group(1).lower()
        contenido = base64.b64decode(datos) if "base64" in cabecera else \
            urllib.parse.unquote_to_bytes(datos)
        ext = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp",
               "image/gif": "gif"}.get(ct, "jpg")
        return contenido, ct, ext
    r = requests.get(url, headers={"User-Agent": BROWSER_UA, "Accept": "image/*,*/*;q=0.8",
                                   "Referer": url},
                     timeout=25, allow_redirects=True)
    r.raise_for_status()
    ct = (r.headers.get("content-type") or "").split(";")[0].strip().lower()
    if not ct.startswith("image/") or len(r.content) < 1000:
        raise RuntimeError(f"la URL no devuelve una imagen ({ct or 'sin tipo'})")
    if len(r.content) > 10 * 1024 * 1024:
        raise RuntimeError("imagen de más de 10 MB")
    ext = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp",
           "image/gif": "gif"}.get(ct, "jpg")
    return r.content, ct, ext


# Umbral mínimo de calidad para las fotos de origen. Por debajo de esto la
# imagen se ve pixelada al ampliarse y las dos plataformas la rechazan
# ("No cumple el tamaño mínimo..."). Exigimos un lado corto de al menos 600 px
# y un lado largo de al menos 1000 px: así, al encuadrarla, casi no hace falta
# agrandarla y no queda borrosa.
FOTO_MIN_LADO_CORTO = 600
FOTO_MIN_LADO_LARGO = 1000


def dim_imagen(contenido):
    """(ancho, alto) de la imagen, o None si no se puede leer."""
    try:
        from PIL import Image
        with Image.open(io.BytesIO(contenido)) as img:
            return img.size
    except Exception:
        return None


def foto_calidad_ok(contenido):
    """True si la imagen tiene resolución suficiente para publicar."""
    dim = dim_imagen(contenido)
    if not dim:
        return False
    w, h = dim
    return min(w, h) >= FOTO_MIN_LADO_CORTO and max(w, h) >= FOTO_MIN_LADO_LARGO


def firma_imagen(contenido):
    """Huella del CONTENIDO visual de la imagen (miniatura 16x16 en grises).
    Sirve para detectar fotos repetidas aunque vengan con URLs distintas o en
    tamaños distintos, y así no subir la misma foto 2, 3 o 4 veces."""
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(contenido)).convert("L").resize((16, 16), Image.LANCZOS)
        return bytes(img.getdata())
    except Exception:
        # Si no se puede decodificar, cae al hash de los bytes crudos
        import hashlib
        return hashlib.md5(contenido).digest()


# Cuánto se permite agrandar una foto respecto de su tamaño real. Más que esto
# se empieza a ver borrosa, así que es el tope aunque el usuario pida más zoom.
FOTO_MAX_UPSCALE = 1.35


def _esquinas_oscuras(img):
    """True si al menos 3 de las 4 esquinas son casi negras: señal de que la
    foto tiene fondo negro (transparencia que se aplanó mal en su momento)."""
    try:
        w, h = img.size
        pts = [(1, 1), (w - 2, 1), (1, h - 2), (w - 2, h - 2)]
        return sum(1 for p in pts if sum(img.getpixel(p)[:3]) < 60) >= 3
    except Exception:
        return False


def _ratio_blanco(img):
    """Fracción de la imagen que quedó casi blanca (para detectar si el flood
    fill se comió el producto)."""
    try:
        chico = img.resize((40, 40))
        px = list(chico.getdata())
        blancos = sum(1 for p in px if sum(p[:3]) > 720)
        return blancos / max(1, len(px))
    except Exception:
        return 0.0


def aclarar_fondo_negro(contenido):
    """Convierte un fondo negro sólido en blanco (flood fill desde las 4
    esquinas). Sirve para recuperar fotos que quedaron con fondo negro, PERO
    solo si el producto es claro: si el flood fill se termina comiendo casi toda
    la imagen (producto oscuro sobre negro), se aborta para no borrarla.
    Devuelve bytes JPEG, o None si no hacía falta / no era seguro."""
    try:
        from PIL import Image, ImageDraw
        img = Image.open(io.BytesIO(contenido)).convert("RGB")
        if not _esquinas_oscuras(img):
            return None
        blanco_antes = _ratio_blanco(img)
        w, h = img.size
        for c in [(0, 0), (w - 1, 0), (0, h - 1), (w - 1, h - 1)]:
            try:
                ImageDraw.floodfill(img, c, (255, 255, 255), thresh=45)
            except Exception:
                pass
        # Freno de seguridad: si el relleno dejó casi todo blanco, el producto
        # era oscuro y se lo comió -> abortar (mejor dejar la foto como estaba).
        if _ratio_blanco(img) - blanco_antes > 0.5 or _ratio_blanco(img) > 0.9:
            return None
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=95)
        return out.getvalue()
    except Exception:
        return None


def encuadrar(contenido, W=1740, H=1170, fill=0.92):
    """Encuadra la imagen en un lienzo WxH con fondo blanco, sin deformar.
    Por defecto 1740x1170 (3:2, estándar Tiendanube); para ML se usa 1200x1200.

    `fill` (0.5 a 1.0) es cuánto del lienzo ocupa el producto: es el 'zoom' que
    elige el usuario. Se respeta salvo que agrandar tanto arruine la calidad:
    nunca se supera FOTO_MAX_UPSCALE respecto del tamaño original."""
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(contenido))
        # Si la imagen tiene transparencia (PNG/WebP), la aplanamos sobre BLANCO.
        # Con un simple convert("RGB"), PIL rellena lo transparente con NEGRO.
        if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
            fondo = Image.new("RGBA", img.size, (255, 255, 255, 255))
            img = Image.alpha_composite(fondo, img.convert("RGBA")).convert("RGB")
        else:
            img = img.convert("RGB")
        fill = max(0.5, min(float(fill or 0.92), 1.0))
        escala = min(W * fill / img.width, H * fill / img.height)
        escala = min(escala, FOTO_MAX_UPSCALE)  # no agrandar más de la cuenta
        nw, nh = max(1, round(img.width * escala)), max(1, round(img.height * escala))
        img = img.resize((nw, nh), Image.LANCZOS)
        lienzo = Image.new("RGB", (W, H), (255, 255, 255))
        lienzo.paste(img, ((W - nw) // 2, (H - nh) // 2))
        out = io.BytesIO()
        lienzo.save(out, format="JPEG", quality=90)
        return out.getvalue(), "image/jpeg", "jpg"
    except Exception:
        return None


def encuadrar_1740x1170(contenido, fill=0.92):
    """Compatibilidad: encuadre estándar de Tiendanube (3:2)."""
    return encuadrar(contenido, 1740, 1170, fill)


def ml_subir_foto(token, url, vistas=None, fill=0.92):
    """Descarga la imagen, verifica su calidad, descarta las repetidas (misma
    foto ya vista), la reencuadra a un cuadrado 1200x1200 (fondo blanco, así
    cumple la proporción que exige ML) y la sube al hosting de fotos de ML.
    Devuelve el picture id, o None si es de baja calidad, repetida o falló."""
    try:
        contenido, ct, ext = descargar_imagen(url)
        if not foto_calidad_ok(contenido):
            return None
        if vistas is not None:
            firma = firma_imagen(contenido)
            if firma in vistas:
                return None
            vistas.add(firma)
        enc = encuadrar(contenido, 1200, 1200, fill)
        if enc:
            contenido, ct, ext = enc
        r = requests.post(ML_BASE + "/pictures/items/upload",
                          headers={"Authorization": f"Bearer {token}"},
                          files={"file": (f"foto.{ext}", contenido, ct)}, timeout=60)
        if r.status_code in (200, 201):
            return r.json().get("id")
    except Exception:
        pass
    return None


def sku_raiz(sku):
    if not sku:
        return None
    return str(sku).strip().split(".")[0].upper() or None


def gtins_de_item_ml(item):
    """Códigos de barras (GTIN/EAN) del item y sus variaciones, para cruzar
    con el ERP cuando la publicación no tiene SKU cargado."""
    gtins = set()
    for a in item.get("attributes") or []:
        if a.get("id") in ("GTIN", "EAN") and a.get("value_name"):
            gtins.add(str(a["value_name"]).strip())
    for v in item.get("variations") or []:
        for a in v.get("attributes") or []:
            if a.get("id") in ("GTIN", "EAN") and a.get("value_name"):
                gtins.add(str(a["value_name"]).strip())
        for a in v.get("attribute_combinations") or []:
            if a.get("id") in ("GTIN", "EAN") and a.get("value_name"):
                gtins.add(str(a["value_name"]).strip())
    return gtins


def ml_listar_ids(token, user_id, status):
    """Todos los IDs de items del vendedor. Usa scan (sin tope) y cae a
    paginación por offset (tope 1000) si scan no funciona o no avanza."""
    ids, vistos, scroll = [], set(), None
    scan_completo = False
    try:
        for _ in range(300):
            params = {"status": status, "limit": 100, "search_type": "scan"}
            if scroll:
                params["scroll_id"] = scroll
            data = ml_get(f"/users/{user_id}/items/search", token, params)
            res = data.get("results", [])
            total = data.get("paging", {}).get("total", 0)
            if not res:
                scan_completo = True
                break
            nuevos = [x for x in res if x not in vistos]
            ids += nuevos
            vistos.update(nuevos)
            if total and len(ids) >= total:
                scan_completo = True
                break
            nuevo_scroll = data.get("scroll_id")
            if not nuevo_scroll or not nuevos:
                # scan sin cursor o repitiendo la misma página: no sirve
                break
            scroll = nuevo_scroll
    except requests.HTTPError:
        pass
    if scan_completo:
        return ids
    # Fallback: offset clásico (tope 1000 por estado)
    ids, offset = [], 0
    while True:
        data = ml_get(f"/users/{user_id}/items/search", token,
                      {"status": status, "limit": 100, "offset": offset})
        res = data.get("results", [])
        ids += res
        total = data.get("paging", {}).get("total", 0)
        offset += 100
        if not res or offset >= min(total, 1000):
            break
    return ids


def skus_de_item_ml(item):
    """Junta los SKUs del item y de sus variaciones."""
    skus = set()
    if item.get("seller_custom_field"):
        skus.add(str(item["seller_custom_field"]).strip())
    for a in item.get("attributes") or []:
        if a.get("id") == "SELLER_SKU" and a.get("value_name"):
            skus.add(str(a["value_name"]).strip())
    for v in item.get("variations") or []:
        if v.get("seller_custom_field"):
            skus.add(str(v["seller_custom_field"]).strip())
        for a in v.get("attributes") or []:
            if a.get("id") == "SELLER_SKU" and a.get("value_name"):
                skus.add(str(a["value_name"]).strip())
    return skus


# ---------------------------------------------------------------- helpers TN

def tn_headers(token):
    return {"Authentication": f"bearer {token}",
            "User-Agent": TN_UA,
            "Content-Type": "application/json"}


def tn_nombre(name):
    """El name de Tiendanube es un dict por idioma."""
    if isinstance(name, dict):
        return name.get("es") or next(iter(name.values()), "")
    return name or ""


# ------------------------------------------------------------- stock (Hansa)

def parse_num(v):
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        try:
            return float(s.replace(".", "").replace(",", "."))
        except ValueError:
            return 0.0


def parse_stock_rows(rows):
    """Filas crudas de la planilla -> productos agrupados por SKU raíz.

    Columnas: código, descripción, código de barras, y varias columnas de stock
    por depósito. La plataforma sólo debe ver el stock del depósito LOYOLA: se
    detecta esa columna por su encabezado y se usa únicamente esa. Si no se
    encuentra (planilla con otro formato), se cae a sumar las columnas 3-5.
    """
    productos = {}
    idx_loyola = None
    for row in rows:
        vals = [str(c).strip() if c is not None else "" for c in row]
        codigo = vals[0] if vals else ""
        # Detectar la columna del depósito LOYOLA en el encabezado (sólo entre
        # las columnas de stock, de la 3 en adelante, para no confundir con un
        # nombre de producto que diga "loyola").
        if idx_loyola is None:
            for i in range(3, len(vals)):
                if "loyola" in vals[i].lower():
                    idx_loyola = i
                    break
        if not codigo or "sincronización" in codigo.lower() or codigo.lower().startswith("código"):
            continue
        nombre  = str((row[1] if len(row) > 1 else "") or "").strip()
        barcode = str((row[2] if len(row) > 2 else "") or "").strip()
        if idx_loyola is not None:
            stock = parse_num(row[idx_loyola]) if len(row) > idx_loyola else 0
        else:
            stock = sum(parse_num(row[i]) for i in (3, 4, 5) if len(row) > i)
        raiz, _, sufijo = codigo.partition(".")
        raiz = raiz.strip().upper()
        p = productos.setdefault(raiz, {
            "sku_raiz": raiz, "nombre": "", "variantes": [], "stock_total": 0.0,
        })
        if nombre and not p["nombre"]:
            p["nombre"] = nombre
        p["variantes"].append({
            "sku": codigo, "sufijo": sufijo.strip(), "barcode": barcode,
            "stock": stock, "nombre": nombre,
        })
        p["stock_total"] += stock
    for p in productos.values():
        p["tiene_variantes"] = any(v["sufijo"] for v in p["variantes"])
    return productos


def limpiar_codigo(v):
    """'2027.0' -> '2027' (números que la planilla formatea como float)."""
    s = str(v or "").strip()
    if s.endswith(".0") and s[:-2].replace(".", "", 1).isdigit():
        s = s[:-2]
    return s


def parse_maestro_alt(rows):
    """Solapa 'Maestro Items' -> {SKU: {alt, grupos, ml_ids, tn_id}}.

    alt = Código Alternativo (código del fabricante); grupos = códigos de
    clasificación (Grupos Display); ml_ids / tn_id = vínculos que Hansa ya
    tiene con las publicaciones de cada plataforma.
    """
    maestro, idx = {}, None
    for row in rows:
        vals = [str(c).strip() if c is not None else "" for c in row]
        if idx is None:
            if "Código Alternativo" in vals:
                # Código de barras (GTIN/EAN): por nombre de encabezado, y si no
                # aparece, cae a la columna I (índice 8), donde lo tiene la planilla.
                bc = next((i for i, v in enumerate(vals)
                           if any(k in v.lower() for k in
                                  ("barra", "barcode", "ean", "gtin", "cod. barra"))), None)
                if bc is None and len(vals) > 8:
                    bc = 8
                idx = {
                    "alt": vals.index("Código Alternativo"),
                    "grupos": vals.index("Grupos Display") if "Grupos Display" in vals else None,
                    "ml": vals.index("ML IDs") if "ML IDs" in vals else None,
                    "tn": vals.index("TN ID") if "TN ID" in vals else None,
                    "barcode": bc,
                    # Lista de precios de venta (ej: "Precio Lista 1") si la
                    # planilla la incorpora; el Precio Costo no cuenta
                    "precio": next((i for i, v in enumerate(vals)
                                    if v.lower().startswith("precio")
                                    and "costo" not in v.lower()), None),
                }
            continue
        cod = vals[0] if vals else ""
        if not cod:
            continue

        def celda(i):
            return vals[i] if i is not None and len(vals) > i else ""

        ml_ids = [x.strip() for x in celda(idx["ml"]).split(",")
                  if x.strip().upper().startswith("MLA")]
        tn_raw = celda(idx["tn"])
        tn_id = ""
        if tn_raw.upper().startswith("OK"):
            digitos = re.sub(r"\D", "", tn_raw)
            tn_id = digitos or "ok"
        # El código de barras a veces viene como float ("7791234567890.0")
        barcode = re.sub(r"\D", "", limpiar_codigo(celda(idx.get("barcode"))))
        maestro[cod.upper()] = {
            "alt": limpiar_codigo(celda(idx["alt"])),
            "grupos": [g.strip().upper() for g in celda(idx["grupos"]).split(",") if g.strip()],
            "ml_ids": ml_ids,
            "tn_id": tn_id,
            "barcode": barcode,
            "precio": parse_num(celda(idx["precio"])),
        }
    return maestro


def parse_precios_retail(rows):
    """Solapa 'Precios retail' (Item | Nombre | Unidad | IVA Incl.) ->
    {SKU raíz: precio de venta con IVA}."""
    precios, idx_precio = {}, None
    for row in rows:
        vals = [str(c).strip() if c is not None else "" for c in row]
        if idx_precio is None:
            if any("iva" in v.lower() for v in vals):
                idx_precio = next(i for i, v in enumerate(vals) if "iva" in v.lower())
            continue
        cod = vals[0] if vals else ""
        p = parse_num(vals[idx_precio]) if len(vals) > idx_precio else 0
        if cod and p > 0:
            precios[cod.upper()] = p
    return precios


def aplicar_alt(productos, maestro, precios=None):
    """Vuelca los datos del Maestro (alt, grupos, vínculos ML/TN) en cada
    producto y variante."""
    for p in productos.values():
        grupos, ml_ids, tn_id = set(), set(), ""
        for v in p["variantes"]:
            m = maestro.get(v["sku"].upper()) or {}
            v["alt"] = m.get("alt", "")
            v["barcode_hansa"] = m.get("barcode", "")
            grupos.update(m.get("grupos") or [])
            ml_ids.update(m.get("ml_ids") or [])
            tn_id = tn_id or m.get("tn_id", "")
        raiz = maestro.get(p["sku_raiz"]) or {}
        grupos.update(raiz.get("grupos") or [])
        ml_ids.update(raiz.get("ml_ids") or [])
        p["alt"] = raiz.get("alt", "") or next(
            (v["alt"] for v in p["variantes"] if v.get("alt")), "")
        # Código de barras: el del SKU raíz, o el de la primera variante que lo tenga
        p["barcode"] = raiz.get("barcode", "") or next(
            (v["barcode_hansa"] for v in p["variantes"] if v.get("barcode_hansa")), "")
        p["grupos"] = sorted(grupos)
        p["ml_ids_hansa"] = sorted(ml_ids)
        p["tn_id_hansa"] = tn_id or raiz.get("tn_id", "")
        # Precio: primero la lista retail (por SKU raíz — las variantes
        # comparten precio), después alguna columna de precio del Maestro
        p["precio"] = (precios or {}).get(p["sku_raiz"], 0) or raiz.get("precio") or max(
            (maestro.get(v["sku"].upper(), {}).get("precio", 0) for v in p["variantes"]),
            default=0)
    return productos


def fetch_stock_xlsx(sheet_id):
    """Baja el workbook completo: solapa Stock + Maestro Items (alt codes)."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    r = requests.get(url, timeout=60, allow_redirects=True)
    if r.status_code != 200 or "text/html" in r.headers.get("content-type", ""):
        raise RuntimeError(
            "No se pudo leer la planilla. Verificá que esté compartida como "
            "'Cualquier persona con el enlace: Lector', o subí el archivo manualmente."
        )
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    stock_rows, alt_map, precios = None, {}, {}
    for ws in wb.worksheets:
        titulo = ws.title.strip().lower()
        if titulo == "stock" or (stock_rows is None and "stock" in titulo):
            stock_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        elif "maestro" in titulo:
            alt_map = parse_maestro_alt(ws.iter_rows(values_only=True))
        elif "precio" in titulo:
            precios = parse_precios_retail(ws.iter_rows(values_only=True))
    if stock_rows is None:
        stock_rows = [list(row) for row in wb.worksheets[0].iter_rows(values_only=True)]
    return stock_rows, alt_map, precios


def fetch_stock_sheet(sheet_id, gid):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    r = requests.get(url, timeout=30, allow_redirects=True)
    ct = r.headers.get("content-type", "")
    if r.status_code != 200 or "text/html" in ct:
        raise RuntimeError(
            "No se pudo leer la planilla. Verificá que esté compartida como "
            "'Cualquier persona con el enlace: Lector', o subí el CSV manualmente."
        )
    text = r.content.decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


@sync_bp.route("/stock")
def get_stock():
    sheet_id = request.args.get("sheet_id", SHEET_ID)
    gid      = request.args.get("gid", SHEET_GID)
    try:
        try:
            rows, alt_map, precios = fetch_stock_xlsx(sheet_id)
        except Exception:
            rows, alt_map, precios = fetch_stock_sheet(sheet_id, gid), {}, {}
        productos = aplicar_alt(parse_stock_rows(rows), alt_map, precios)
        try:
            with open(STOCK_CACHE_PATH, "w") as f:
                json.dump(productos, f)
        except OSError:
            pass
        return jsonify({"productos": productos, "total_skus": len(productos), "fuente": "sheet"})
    except Exception as e:
        # Fallback al último parse guardado
        if os.path.exists(STOCK_CACHE_PATH):
            with open(STOCK_CACHE_PATH) as f:
                productos = json.load(f)
            return jsonify({"productos": productos, "total_skus": len(productos),
                            "fuente": "cache", "warning": str(e)})
        return jsonify({"error": str(e)}), 502


@sync_bp.route("/stock", methods=["POST"])
def upload_stock():
    """Fallback: subir el CSV exportado de la planilla a mano."""
    if "file" not in request.files:
        return jsonify({"error": "No se encontró el archivo"}), 400
    f = request.files["file"]
    try:
        name = (f.filename or "").lower()
        alt_map, precios = {}, {}
        if name.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            rows = [list(r) for r in wb.worksheets[0].iter_rows(values_only=True)]
            for ws in wb.worksheets:
                titulo = ws.title.strip().lower()
                if "maestro" in titulo:
                    alt_map = parse_maestro_alt(ws.iter_rows(values_only=True))
                elif "precio" in titulo:
                    precios = parse_precios_retail(ws.iter_rows(values_only=True))
        else:
            text = f.read().decode("utf-8-sig", errors="replace")
            rows = list(csv.reader(io.StringIO(text)))
        productos = aplicar_alt(parse_stock_rows(rows), alt_map, precios)
        try:
            with open(STOCK_CACHE_PATH, "w") as cf:
                json.dump(productos, cf)
        except OSError:
            pass
        return jsonify({"productos": productos, "total_skus": len(productos), "fuente": "upload"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ------------------------------------------------------ equivalencias Hansa

# La planilla de equivalencias tiene celdas con mojibake (UTF-8 leído mal al
# importar desde Hansa): '√±' es 'ñ', '√©' es 'é', etc. Se corrige al leer.
_MOJIBAKE = {"√±": "ñ", "√ë": "Ñ", "√°": "á", "√©": "é", "√≠": "í",
             "√≥": "ó", "√∫": "ú", "√º": "ü", "¬∞": "°"}


def _celda(row, i):
    if len(row) <= i or row[i] is None:
        return ""
    s = str(row[i]).strip()
    for feo, bien in _MOJIBAKE.items():
        if feo in s:
            s = s.replace(feo, bien)
    return s


@sync_bp.route("/equivalencias")
def equivalencias():
    """Lee la planilla de equivalencias: variantes (código -> nombre y tipo),
    categorías TN por tipo de producto, marca por SKU madre y URL del
    fabricante (solapa SLUGS)."""
    sheet_id = request.args.get("sheet_id", EQUIV_SHEET_ID)
    try:
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        r = requests.get(url, timeout=30, allow_redirects=True)
        if r.status_code != 200 or "text/html" in r.headers.get("content-type", ""):
            raise RuntimeError(
                "No se pudo leer la planilla de equivalencias. Verificá que esté "
                "compartida como 'Cualquier persona con el enlace: Lector'."
            )
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)

        def hoja(nombre):
            for ws in wb.worksheets:
                if ws.title.strip().lower() == nombre:
                    return ws
            return None

        data = {"variantes": {}, "categorias_tn": {}, "marcas": {}, "slugs": {}}

        ws = hoja("variantes")
        if ws:
            for row in list(ws.iter_rows(values_only=True))[1:]:
                cod = _celda(row, 0)
                if cod:
                    data["variantes"][cod.upper()] = {
                        "nombre": _celda(row, 1) or cod,
                        "tipo": _celda(row, 2).upper() or "COLOR",
                    }

        ws = hoja("categorias tn")
        if ws:
            for row in list(ws.iter_rows(values_only=True))[1:]:
                nombre = _celda(row, 1)
                cats = []
                for i in (2, 4, 6):
                    cid = _celda(row, i)
                    if cid:
                        try:
                            cats.append({"id": int(float(cid)), "nombre": _celda(row, i + 1)})
                        except ValueError:
                            pass
                if nombre and cats:
                    data["categorias_tn"][nombre.lower()] = cats
                    cod = _celda(row, 0)
                    if cod:
                        data["categorias_tn"][cod.upper()] = cats

        ws = hoja("publicar")
        if ws:
            for row in list(ws.iter_rows(values_only=True))[1:]:
                sku_madre, marca = _celda(row, 1), _celda(row, 7)
                if sku_madre and marca:
                    data["marcas"][sku_madre.upper()] = marca

        # Clasificaciones Hansa (código -> nombre, tipo y —para tipo PROD—
        # peso y dimensiones del paquete de envío). tipo MAR = marca.
        data["clasificaciones"] = {}
        ws = hoja("clasificaciones")
        if ws:
            for row in list(ws.iter_rows(values_only=True))[1:]:
                cod = _celda(row, 0)
                if not cod:
                    continue
                entry = {"nombre": _celda(row, 1) or cod, "tipo": _celda(row, 2).upper()}
                peso  = parse_num(_celda(row, 3))
                largo = parse_num(_celda(row, 4))
                alto  = parse_num(_celda(row, 5))
                ancho = parse_num(_celda(row, 6))
                if peso or largo or alto or ancho:
                    entry["dim"] = {"peso_kg": peso, "largo": largo, "alto": alto, "ancho": ancho}
                data["clasificaciones"][cod.upper()] = entry

        ws = hoja("slugs")
        if ws:
            for row in list(ws.iter_rows(values_only=True))[1:]:
                sku, slug, base = _celda(row, 0), _celda(row, 2), _celda(row, 3)
                if sku and base:
                    full = base + slug if base.endswith("/") else (base.rstrip("/") + "/" + slug if slug else base)
                    data["slugs"][sku.upper()] = {"url": full, "notas": _celda(row, 4)}

        try:
            with open(EQUIV_CACHE_PATH, "w") as f:
                json.dump(data, f)
        except OSError:
            pass
        return jsonify(data)
    except Exception as e:
        if os.path.exists(EQUIV_CACHE_PATH):
            with open(EQUIV_CACHE_PATH) as f:
                data = json.load(f)
            data["warning"] = str(e)
            return jsonify(data)
        return jsonify({"error": str(e)}), 502


# ------------------------------------------------------------- Mercado Libre

@sync_bp.route("/ml")
def get_ml():
    token   = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    user_id = request.args.get("user_id", os.environ.get("ML_USER_ID", "246901020"))
    if not token:
        return jsonify({"error": "Falta el token de Mercado Libre"}), 400
    try:
        ids = []
        for status in ("active", "paused", "under_review", "inactive"):
            try:
                ids += ml_listar_ids(token, user_id, status)
            except requests.HTTPError:
                continue
        items_por_raiz = {}
        sin_sku = []
        for i in range(0, len(ids), 20):
            chunk = ids[i:i + 20]
            # include_attributes=all: sin esto ML recorta atributos (incluido
            # el SELLER_SKU de las variaciones) en las consultas multiget
            details = ml_get("/items", token, {"ids": ",".join(chunk), "include_attributes": "all"})
            for x in details:
                if x.get("code") != 200:
                    continue
                item = x["body"]
                resumen = {"id": item.get("id"), "title": item.get("title"),
                           "status": item.get("status"), "permalink": item.get("permalink"),
                           "price": item.get("price"),
                           "gtins": sorted(gtins_de_item_ml(item))}
                raices = {sku_raiz(s) for s in skus_de_item_ml(item)} - {None}
                if not raices:
                    sin_sku.append(resumen)
                for r in raices:
                    items_por_raiz.setdefault(r, []).append(resumen)
        return jsonify({"items_por_raiz": items_por_raiz, "sin_sku": sin_sku,
                        "total_items": len(ids)})
    except requests.HTTPError as e:
        return jsonify({"error": f"Mercado Libre: {e.response.status_code} {e.response.text[:300]}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@sync_bp.route("/ml/ids")
def ml_ids():
    """Solo los IDs de publicaciones (rápido). El detalle se pide por tandas
    con /ml/detalles para no exceder el timeout del servidor."""
    token   = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    user_id = request.args.get("user_id", os.environ.get("ML_USER_ID", "246901020"))
    if not token:
        return jsonify({"error": "Falta el token de Mercado Libre"}), 400
    try:
        ids = []
        # under_review e inactive también cuentan como "ya publicado": si no
        # se leen, esos items figuran como faltantes aunque existan
        for status in ("active", "paused", "under_review", "inactive"):
            try:
                ids += ml_listar_ids(token, user_id, status)
            except requests.HTTPError:
                continue  # algún estado puede no estar habilitado para la cuenta
        return jsonify({"ids": ids})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@sync_bp.route("/ml/detalles", methods=["POST"])
def ml_detalles():
    """Detalle (SKUs, GTINs, estado) de hasta 500 publicaciones por llamada."""
    body  = request.json or {}
    token = body.get("token", os.environ.get("ML_TOKEN", ""))
    ids   = body.get("ids") or []
    if not token or not ids:
        return jsonify({"error": "Faltan token o ids"}), 400
    ids = ids[:500]
    try:
        out = []
        for i in range(0, len(ids), 20):
            chunk = ids[i:i + 20]
            # include_attributes=all: sin esto ML recorta atributos (incluido
            # el SELLER_SKU de las variaciones) en las consultas multiget
            details = ml_get("/items", token, {"ids": ",".join(chunk), "include_attributes": "all"})
            for x in details:
                if x.get("code") != 200:
                    continue
                item = x["body"]
                resumen = {"id": item.get("id"), "title": item.get("title"),
                           "status": item.get("status"), "permalink": item.get("permalink"),
                           "price": item.get("price"),
                           "gtins": sorted(gtins_de_item_ml(item))}
                raices = sorted({sku_raiz(s) for s in skus_de_item_ml(item)} - {None})
                out.append({"resumen": resumen, "raices": raices})
        return jsonify({"items": out})
    except requests.HTTPError as e:
        return jsonify({"error": f"Mercado Libre: {e.response.status_code} {e.response.text[:300]}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@sync_bp.route("/ml/fotos")
def ml_fotos():
    """Fotos de una publicación existente de ML (para reutilizarlas al
    publicar el mismo producto en Tiendanube)."""
    token   = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    item_id = request.args.get("item_id", "")
    if not token or not item_id:
        return jsonify({"fotos": [], "error": "Faltan parámetros"})
    try:
        item = ml_get(f"/items/{item_id}", token)
        return jsonify({"fotos": [p.get("secure_url") or p.get("url")
                                  for p in item.get("pictures") or []
                                  if p.get("secure_url") or p.get("url")]})
    except Exception as e:
        return jsonify({"fotos": [], "error": str(e)})


@sync_bp.route("/ml/debug")
def ml_debug():
    """Item completo de ML (con todos los atributos) para diagnosticar por
    qué una publicación no cruza con el ERP."""
    token   = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    item_id = request.args.get("item_id", "")
    if not token or not item_id:
        return jsonify({"error": "Faltan token o item_id"}), 400
    try:
        item = ml_get(f"/items/{item_id}", token, {"include_attributes": "all"})
        user_id = request.args.get("user_id", os.environ.get("ML_USER_ID", "246901020"))
        # Motivo de baja / estado de salud (por qué ML la dio de baja)
        salud = None
        try:
            salud = ml_get(f"/items/{item_id}/health", token)
        except Exception:
            pass
        return jsonify({
            "id": item.get("id"),
            "title": item.get("title"),
            "status": item.get("status"),
            "sub_status": item.get("sub_status"),
            "tags": item.get("tags"),
            "health": item.get("health"),
            "health_detalle": salud,
            "catalog_listing": item.get("catalog_listing"),
            "catalog_product_id": item.get("catalog_product_id"),
            "seller_id": item.get("seller_id"),
            "es_de_esta_cuenta": str(item.get("seller_id")) == str(user_id),
            "cuenta_configurada": user_id,
            "seller_custom_field": item.get("seller_custom_field"),
            "skus_detectados": sorted(skus_de_item_ml(item)),
            "gtins_detectados": sorted(gtins_de_item_ml(item)),
            "attributes": [{"id": a.get("id"), "value_name": a.get("value_name")}
                           for a in item.get("attributes") or []],
            "variations": [{
                "id": v.get("id"),
                "seller_custom_field": v.get("seller_custom_field"),
                "attributes": [{"id": a.get("id"), "value_name": a.get("value_name")}
                               for a in v.get("attributes") or []],
                "attribute_combinations": [{"id": a.get("id"), "value_name": a.get("value_name")}
                                           for a in v.get("attribute_combinations") or []],
            } for v in item.get("variations") or []],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@sync_bp.route("/ml/token", methods=["POST"])
def ml_token():
    """Canjea el code de OAuth (o un refresh_token) por un access token de ML."""
    body   = request.json or {}
    app_id = body.get("app_id", "").strip()
    secret = body.get("secret", "").strip()
    if not app_id or not secret:
        return jsonify({"error": "Faltan App ID o Client Secret de Mercado Libre"}), 400
    payload = {"client_id": app_id, "client_secret": secret}
    if body.get("code"):
        payload.update({"grant_type": "authorization_code", "code": body["code"],
                        "redirect_uri": body.get("redirect_uri", "")})
    elif body.get("refresh_token"):
        payload.update({"grant_type": "refresh_token", "refresh_token": body["refresh_token"]})
    else:
        return jsonify({"error": "Falta el code o el refresh_token"}), 400
    try:
        r = requests.post(ML_BASE + "/oauth/token", data=payload,
                          headers={"Accept": "application/json"}, timeout=20)
    except requests.RequestException as e:
        return jsonify({"error": f"No se pudo contactar a Mercado Libre: {e}"}), 502
    if r.status_code != 200:
        try:
            detail = r.json()
        except ValueError:
            detail = r.text[:300]
        return jsonify({"error": "Mercado Libre rechazó la autorización", "detalle": detail}), 502
    d = r.json()
    return jsonify({"access_token": d.get("access_token"),
                    "refresh_token": d.get("refresh_token"),
                    "expires_in": d.get("expires_in"),
                    "user_id": d.get("user_id")})


def _variantes_query(q):
    """Genera consultas de categoría, de la más específica a la más genérica,
    para maximizar la chance de que ML prediga algo. Ej.:
    'Cambio shifter Shimano CUES SL-U4010-9R (9 vel) Rapidfire' ->
    [título completo, sin paréntesis, primeras 4/3/2 palabras]."""
    q = (q or "").strip()
    variantes = []
    def agregar(v):
        v = re.sub(r"\s+", " ", (v or "").strip())
        if v and v.lower() not in {x.lower() for x in variantes}:
            variantes.append(v)
    agregar(q)
    agregar(re.sub(r"\([^)]*\)", "", q))          # sin paréntesis
    palabras = re.sub(r"\([^)]*\)", "", q).split()
    for n in (4, 3, 2):
        if len(palabras) >= n:
            agregar(" ".join(palabras[:n]))
    # Variante "palabras clave": saca conectores, números y códigos de modelo
    # (ej. 'Casco de moto Abierto LS2 562 Airflow 2 solid' -> 'casco moto abierto').
    _stop = {"de", "la", "el", "los", "las", "para", "con", "sin", "y", "o", "a",
             "un", "una", "del", "al", "por"}
    claves = [w for w in re.sub(r"\([^)]*\)", "", q).lower().split()
              if w not in _stop and not re.search(r"\d", w) and len(w) > 2]
    for n in (3, 2, 1):
        if len(claves) >= n:
            agregar(" ".join(claves[:n]))
    return variantes


@sync_bp.route("/ml/categoria")
def ml_categoria():
    """Predicción de categoría de ML para un título. Prueba varias consultas
    (de específica a genérica) y devuelve la primera que arroje resultados, así
    casi nunca hace falta cargarla a mano."""
    token = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    q = request.args.get("q", "")
    if not q:
        return jsonify({"error": "Falta q"}), 400
    ultimo_error = None
    for variante in _variantes_query(q):
        try:
            data = ml_get("/sites/MLA/domain_discovery/search", token,
                          {"q": variante, "limit": 3})
            if data:
                return jsonify({"predicciones": [
                    {"category_id": d.get("category_id"), "category_name": d.get("category_name"),
                     "domain_name": d.get("domain_name")}
                    for d in data
                ], "query_usada": variante})
        except Exception as e:
            ultimo_error = str(e)
    return jsonify({"predicciones": [], "error": ultimo_error})


def _categoria_info(token, cat):
    """Nombre y ruta de una categoría por su ID, o None si no existe."""
    try:
        data = ml_get(f"/categories/{cat}", token, {})
        nombre = data.get("name") if isinstance(data, dict) else None
        if not nombre:
            return None
        ruta = " > ".join(p.get("name", "") for p in (data.get("path_from_root") or []))
        return {"name": nombre, "ruta": ruta}
    except Exception:
        return None


@sync_bp.route("/ml/categoria/nombre")
def ml_categoria_nombre():
    """Nombre de una categoría de ML a partir de su ID (para validar la carga
    manual cuando la predicción no encontró categoría)."""
    token = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    cat   = request.args.get("category_id", "").strip()
    if not cat:
        return jsonify({"error": "Falta category_id"}), 400
    info = _categoria_info(token, cat)
    if not info:
        return jsonify({"error": "categoría inexistente"})
    return jsonify(info)


@sync_bp.route("/ml/categoria/resolver")
def ml_categoria_resolver():
    """Resuelve la categoría a partir de un valor que puede ser:
    - un ID de categoría (MLA412445), o
    - el número/ID/link de una PUBLICACIÓN de ML similar (MLA3735766706,
      3735766706, o la URL): en ese caso se toma la categoría de esa publicación.
    Así el usuario puede 'copiar' la categoría de un producto parecido."""
    token = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    valor = (request.args.get("valor", "") or "").strip().upper()
    m = re.search(r"MLA-?(\d+)", valor) or re.search(r"(\d{5,})", valor)
    if not m:
        return jsonify({"error": "No reconocí un ID. Pegá un ID de categoría (MLA...) o el número/link de una publicación de ML."})
    digitos = re.sub(r"\D", "", m.group(0))
    idc = "MLA" + digitos
    # Las categorías tienen pocos dígitos; las publicaciones, muchos (>=8).
    # Se prueba en el orden más probable y, si falla, al revés.
    modos = ["item", "categoria"] if len(digitos) >= 8 else ["categoria", "item"]
    for modo in modos:
        if modo == "categoria":
            info = _categoria_info(token, idc)
            if info:
                return jsonify({"category_id": idc, "name": info["name"],
                                "ruta": info["ruta"], "via": "categoría"})
        else:
            try:
                item = ml_get(f"/items/{idc}", token, {"attributes": "category_id,title"})
                cat = item.get("category_id") if isinstance(item, dict) else None
                if cat:
                    info = _categoria_info(token, cat) or {"name": cat, "ruta": ""}
                    return jsonify({"category_id": cat, "name": info["name"],
                                    "ruta": info["ruta"], "via": "publicación",
                                    "titulo_publicacion": item.get("title", "")})
            except Exception:
                pass
    return jsonify({"error": "No pude resolver la categoría con ese valor (¿es una publicación tuya o un ID válido?)."})


# Atributos que la app ya completa sola (marca, impuestos, paquete, SKU, MPN):
# no hace falta pedírselos al usuario aunque la categoría los marque required.
_ATTR_AUTOCOMPLETADOS = {
    "BRAND", "SELLER_SKU", "MPN", "MODEL", "PART_NUMBER", "GTIN", "EMPTY_GTIN_REASON",
    "VALUE_ADDED_TAX", "IMPORT_DUTY", "SELLER_PACKAGE_WEIGHT", "SELLER_PACKAGE_LENGTH",
    "SELLER_PACKAGE_WIDTH", "SELLER_PACKAGE_HEIGHT", "COLOR", "SIZE",
}


@sync_bp.route("/ml/atributos")
def ml_atributos():
    """Atributos obligatorios de una categoría de ML que la app no completa
    sola, para que el usuario los cargue en el wizard."""
    token = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    cat   = request.args.get("category_id", "")
    tiene_alt = request.args.get("has_alt", "1") != "0"
    # Si el producto no tiene código alternativo, PART_NUMBER/MPN/MODEL no se
    # autocompletan: hay que pedírselos al usuario
    auto = set(_ATTR_AUTOCOMPLETADOS)
    if not tiene_alt:
        auto -= {"PART_NUMBER", "MPN", "MODEL"}
    if not cat:
        return jsonify({"atributos": []})
    try:
        data = ml_get(f"/categories/{cat}/attributes", token)
        req = []
        gtin_pedido = False
        gtin_reasons = []
        gtin_tags = {}
        _req_tags = ("required", "catalog_required", "conditional_required",
                     "catalog_listing_required")
        for a in data:
            tags = a.get("tags") or {}
            obligatorio = tags.get("required") or tags.get("catalog_required")
            # GTIN (código de barras): se maneja aparte, con su propio campo y
            # el motivo "sin GTIN" cuando el producto no tiene código.
            if a.get("id") == "GTIN":
                gtin_tags = tags
                if any(tags.get(k) for k in _req_tags):
                    gtin_pedido = True
                continue
            if a.get("id") == "EMPTY_GTIN_REASON":
                gtin_reasons = [v.get("name") for v in (a.get("values") or []) if v.get("name")]
                continue
            if not obligatorio or a.get("id") in auto:
                continue
            valores = [v.get("name") for v in (a.get("values") or []) if v.get("name")]
            req.append({
                "id": a.get("id"),
                "nombre": a.get("name"),
                "valores": valores[:60],           # lista para el desplegable
                "permite_otro": tags.get("allow_variations") is None,
                "sugerido": _sugerir_valor(a.get("id"), valores),
            })
        # Si la categoría publica motivos "sin GTIN", es porque exige GTIN
        # (aunque su tag no lo diga claramente): así no se nos escapa.
        if gtin_reasons:
            gtin_pedido = True
        return jsonify({"atributos": req,
                        "gtin": {"pedido": gtin_pedido, "motivos": gtin_reasons,
                                 "tags": gtin_tags}})
    except Exception as e:
        return jsonify({"atributos": [], "error": str(e)})


def _sugerir_valor(attr_id, valores):
    """Valor por defecto razonable para Muvin (tienda de bicis)."""
    if attr_id == "VEHICLE_TYPE":
        for v in valores:
            if v.lower() == "bicicleta":
                return v
    return ""


MIN_FOTO_PX = 500  # ancho mínimo aceptable para no publicar imágenes chicas


def _foto_ml(p):
    """De un objeto picture de ML devuelve (url, (ancho, alto)). Usa la URL
    que ML sirve (no fuerza -O, que no siempre existe y rompía la carga)."""
    url = p.get("secure_url") or p.get("url") or ""
    dim = p.get("max_size") or p.get("size") or ""
    m = re.match(r"(\d+)\s*x\s*(\d+)", str(dim))
    wh = (int(m.group(1)), int(m.group(2))) if m else (0, 0)
    return url, wh


def _bing_images(query, limit=20):
    """Imágenes de la web vía Bing (sin API key). Best-effort."""
    urls = []
    r = requests.get("https://www.bing.com/images/search",
                     params={"q": query, "count": limit, "qft": "+filterui:imagesize-large"},
                     headers={"User-Agent": BROWSER_UA, "Accept-Language": "es-AR,es"},
                     timeout=15)
    r.raise_for_status()
    for pat in (r'murl&quot;:&quot;(https?://[^&]+?)&quot;', r'"murl":"(https?://[^"]+?)"'):
        for m in re.finditer(pat, r.text):
            u = m.group(1).replace("\\/", "/")
            if u not in urls:
                urls.append(u)
        if urls:
            break
    return urls[:limit]


def _ddg_images(query, limit=20):
    """Imágenes de la web vía DuckDuckGo (JSON, sin API key). Best-effort."""
    s = requests.Session()
    s.headers.update({"User-Agent": BROWSER_UA, "Accept-Language": "es-AR,es"})
    r = s.get("https://duckduckgo.com/", params={"q": query}, timeout=15)
    m = (re.search(r'vqd=["\']([\d-]+)["\']', r.text)
         or re.search(r'vqd=([\d-]+)\&', r.text)
         or re.search(r'"vqd":"([\d-]+)"', r.text))
    if not m:
        return []
    rj = s.get("https://duckduckgo.com/i.js",
               params={"l": "ar-es", "o": "json", "q": query, "vqd": m.group(1),
                       "f": "size:Large", "p": "1"},
               headers={"Referer": "https://duckduckgo.com/"}, timeout=15)
    data = rj.json()
    return [it["image"] for it in data.get("results", []) if it.get("image")][:limit]


def _imagenes_web(query, limit=20):
    """Prueba DuckDuckGo y Bing; devuelve (urls, error_str)."""
    for fn in (_ddg_images, _bing_images):
        try:
            urls = fn(query, limit)
            if urls:
                return urls, None
        except Exception as e:
            ultimo = f"{fn.__name__}: {e}"
    return [], locals().get("ultimo")


# Patrones de URL que casi nunca son la foto del producto: logos, íconos de
# redes/pagos, avatares, wishlist, banners, blog, etc. Se descartan de una.
_JUNK_IMG = re.compile(
    r'(logo|icon|sprite|thumb|placeholder|banner|flag|whats?app|facebook|'
    r'instagram|twitter|tiktok|youtube|pinterest|linkedin|telegram|avatar|'
    r'gravatar|user[-_]?pic|profile|wishlist|heart|like|favorit|badge|seal|'
    r'review|rating|estrella|star|cuotas?|mercadopago|banco|payment|pago|'
    r'visa|master|amex|tarjet|envio|shipping|garantia|footer|header|nav[-_/]|'
    r'menu|blog|author|comment|social|share|widget|pixel|tracking|spinner|'
    r'loading|lazy|blank|1x1|spacer|default|no[-_]?image|sin[-_]?imagen)', re.I)


def _aspecto_ok(w, h):
    """La foto de un producto es más o menos cuadrada: descarta banners
    (muy anchos) y tiras verticales (muy altas)."""
    if not w or not h:
        return False
    r = w / h
    return 0.5 <= r <= 2.0


def foto_candidata_ok(contenido):
    """Una candidata sirve si se PODRÁ publicar (mismo umbral que la publicación)
    y tiene proporción sana. Así lo que se muestra en la grilla es exactamente lo
    que después se puede publicar (antes se sugerían fotos que luego se
    descartaban por chicas, dejando la publicación sin imágenes)."""
    dim = dim_imagen(contenido)
    if not dim:
        return False
    w, h = dim
    return foto_calidad_ok(contenido) and _aspecto_ok(w, h)


def _verificar_fotos(urls, limite=8, max_bajar=16):
    """Descarga las candidatas y deja solo las que son fotos de producto de
    verdad (tamaño y proporción razonables). Devuelve [{url, size}]."""
    buenas, bajadas = [], 0
    for u in urls:
        if len(buenas) >= limite or bajadas >= max_bajar:
            break
        if _JUNK_IMG.search(u):
            continue
        bajadas += 1
        try:
            contenido, _, _ = descargar_imagen(u)
            if foto_candidata_ok(contenido):
                w, h = dim_imagen(contenido)
                buenas.append({"url": u, "size": f"{w}x{h}"})
        except Exception:
            continue
    return buenas


def _imagenes_fabricante(url):
    """Imágenes de la página del fabricante (og:image + galería)."""
    try:
        r = requests.get(url, headers={"User-Agent": BROWSER_UA}, timeout=15)
        r.raise_for_status()
        html = r.text[:400000]
        urls = []
        # og:image primero (la principal)
        for pat in (r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)',
                    r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image',
                    r'<link[^>]+rel=["\']image_src["\'][^>]+href=["\']([^"\']+)'):
            for m in re.finditer(pat, html, re.I):
                urls.append(m.group(1))
        # Imágenes de la galería del producto (jpg/png/webp grandes)
        for m in re.finditer(r'(?:src|data-src|data-zoom-image|data-large_image)=["\']'
                             r'(https?://[^"\']+\.(?:jpg|jpeg|png|webp)[^"\']*)', html, re.I):
            urls.append(m.group(1))
        base = re.match(r'(https?://[^/]+)', url)
        norm, vistos = [], set()
        for u in urls:
            if u.startswith("//"):
                u = "https:" + u
            elif u.startswith("/") and base:
                u = base.group(1) + u
            if u.startswith("http") and u not in vistos and not _JUNK_IMG.search(u):
                vistos.add(u)
                norm.append(u)
        return norm[:12]
    except Exception:
        pass
    return []


@sync_bp.route("/fotos")
def buscar_fotos():
    """Candidatas de fotos en máxima resolución: página del fabricante +
    catálogo oficial de ML + publicaciones existentes. Descarta las chicas."""
    token    = request.args.get("token", os.environ.get("ML_TOKEN", ""))
    q        = request.args.get("q", "")
    alt      = request.args.get("alt", "").strip()      # código del fabricante
    marca    = request.args.get("marca", "").strip()
    slug_url = request.args.get("slug_url", "").strip()  # página del fabricante
    if not q:
        return jsonify({"error": "Falta q"}), 400
    fotos, vistos = [], set()

    def agregar(url, wh, fuente, tipo):
        if not url or url in vistos:
            return
        vistos.add(url)
        ancho = wh[0] if wh else 0
        # Descartar las claramente chicas (salvo que no sepamos el tamaño)
        if ancho and ancho < MIN_FOTO_PX:
            return
        fotos.append({"url": url, "size": (f"{wh[0]}x{wh[1]}" if ancho else ""),
                      "chica": bool(ancho and ancho < 900), "fuente": fuente,
                      "tipo": tipo})  # fabricante | catalogo | publicacion

    # Cada fuente es independiente: si una falla, las otras siguen. diag
    # cuenta lo que aportó cada una y registra errores para diagnosticar.
    diag = {}

    # 1) Página del fabricante
    if slug_url:
        try:
            n0 = len(fotos)
            for img in _imagenes_fabricante(slug_url):
                agregar(img, (0, 0), "Página del fabricante", "fabricante")
            diag["fabricante"] = len(fotos) - n0
        except Exception as e:
            diag["fabricante_error"] = str(e)[:200]

    # Nota: ML cerró su API pública de búsqueda (/sites/MLA/search → 403), así
    # que las fotos de catálogo/publicaciones de ML ya no están disponibles.
    # Para productos nuevos se usa la búsqueda con IA (/fotos/ia).
    orden = {"fabricante": 0, "catalogo": 1, "publicacion": 2}
    fotos.sort(key=lambda f: (f["chica"], orden.get(f["tipo"], 3)))
    return jsonify({"fotos": fotos[:40], "diag": diag})


@sync_bp.route("/fotos/ia", methods=["POST"])
def fotos_ia():
    """Busca imágenes reales del producto con IA (web_search + web_fetch):
    identifica el producto exacto y devuelve links directos de sus fotos."""
    body    = request.json or {}
    q       = body.get("q", "")
    alt     = (body.get("alt") or "").strip()
    marca   = (body.get("marca") or "").strip()
    slug    = (body.get("slug_url") or "").strip()
    api_key = body.get("api_key") or os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"fotos": [], "error": "Falta la API key de Anthropic (Configuración)."})
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        modelo = re.sub(r"\d+\s*/\s*\d+(\s*/\s*\d+)*", "", q).strip()
        pedido = (f"Buscá en la web la página del producto EXACTO: marca "
                  f"{marca or '(ver nombre)'}, modelo '{modelo}', código de "
                  f"fabricante {alt or '(sin código)'} (rubro ciclismo). "
                  + (f"Empezá por: {slug}. " if slug else "")
                  + "Devolvé ÚNICAMENTE un array JSON con 1 a 4 URLs de PÁGINAS "
                  "(del sitio oficial de la marca o retailers serios) donde se "
                  "vea ese producto, ordenadas de más a menos confiable, así: "
                  "[\"https://...\", \"https://...\"]. Sin texto fuera del array.")
        # Una sola búsqueda con un modelo rápido; la extracción de imágenes la
        # hace nuestro código (rápido), no la IA fetcheando páginas.
        resp = client.messages.create(
            model=os.environ.get("ANTHROPIC_FOTOS_MODEL", "claude-haiku-4-5"),
            max_tokens=600,
            messages=[{"role": "user", "content": pedido}],
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 3}])
        texto = "\n".join(b.text for b in resp.content
                          if getattr(b, "type", "") == "text")
        m = re.search(r"\[.*\]", texto, re.S)
        paginas = json.loads(m.group(0)) if m else []

        # Reunir URLs candidatas (imágenes directas + imágenes de las páginas)
        vistos, candidatas = set(), []
        for pagina in paginas[:4]:
            if not isinstance(pagina, str) or not pagina.startswith("http"):
                continue
            if re.search(r"\.(jpg|jpeg|png|webp)(\?|$)", pagina, re.I):
                if pagina not in vistos:
                    vistos.add(pagina); candidatas.append(pagina)
                continue
            for img in _imagenes_fabricante(pagina):
                if img not in vistos:
                    vistos.add(img); candidatas.append(img)
        # Verificar de verdad: descargar y quedarse solo con las que son fotos
        # de producto (buena resolución y proporción). Así no aparecen logos,
        # íconos de redes, corazones ni banners.
        fotos = [{"url": f["url"], "size": f["size"], "chica": False,
                  "fuente": "Encontrada por IA", "tipo": "ia"}
                 for f in _verificar_fotos(candidatas, limite=8, max_bajar=16)]
        return jsonify({"fotos": fotos})
    except Exception as e:
        return jsonify({"fotos": [], "error": str(e)})


@sync_bp.route("/publish/ml", methods=["POST"])
def publish_ml():
    body  = request.json or {}
    token = body.get("token", os.environ.get("ML_TOKEN", ""))
    item  = body.get("item")
    descripcion = body.get("descripcion", "")
    if not token or not item:
        return jsonify({"error": "Faltan token o item"}), 400

    # ML exige family_name (agrupador de productos del vendedor) en varias
    # categorías: si no vino, usamos el título
    if not item.get("family_name") and item.get("title"):
        item["family_name"] = str(item["title"])[:60]

    # Re-alojar fotos en el hosting de ML para evitar rechazos por hotlink.
    # Las de baja calidad se descartan (ml_subir_foto devuelve None) para que
    # ML no rechace la publicación por "no cumple el tamaño mínimo".
    mapa_fotos = {}
    fotos_ok = []
    fotos_descartadas = 0
    vistas = set()  # firmas de fotos ya subidas, para no repetir
    for pic in item.get("pictures") or []:
        u = pic.get("source")
        if u:
            pid = ml_subir_foto(token, u, vistas, pic.get("zoom", 0.92))
            if pid:
                mapa_fotos[u] = pid
                pic.clear()
                pic["id"] = pid
                fotos_ok.append(pic)
            else:
                fotos_descartadas += 1
    # Sólo dejamos las fotos que se subieron bien
    if item.get("pictures"):
        item["pictures"] = fotos_ok
    for v in item.get("variations") or []:
        if v.get("picture_ids"):
            v["picture_ids"] = [mapa_fotos[u] for u in v["picture_ids"] if u in mapa_fotos]
    foto_warning = None
    if fotos_descartadas:
        foto_warning = (f"Se descartaron {fotos_descartadas} foto(s) repetidas o "
                        "de baja calidad (muy chicas o borrosas) para evitar el rechazo de ML.")

    try:
        r = ml_post("/items", token, item)
        # En el flujo de "familias" de ML, algunas categorías generan el
        # título automáticamente y rechazan el campo title: reintentar sin él.
        # Si el reintento también falla, se informa SU error (es el real).
        if r.status_code == 400 and "title" in r.text and item.get("family_name"):
            reintento = dict(item)
            reintento.pop("title", None)
            r = ml_post("/items", token, reintento)
    except requests.RequestException as e:
        return jsonify({"error": f"No se pudo contactar a Mercado Libre: {e}"}), 502
    if r.status_code not in (200, 201):
        try:
            detail = r.json()
        except ValueError:
            detail = r.text[:500]
        return jsonify({"error": "Mercado Libre rechazó la publicación", "detalle": detail}), 502
    creado = r.json()
    item_id = creado.get("id")
    desc_warning = None
    if descripcion and item_id:
        rd = ml_post(f"/items/{item_id}/description", token, {"plain_text": descripcion})
        if rd.status_code not in (200, 201):
            desc_warning = f"El item se creó pero la descripción falló: {rd.text[:300]}"

    # Aviso inmediato si ML no la dejó activa (quedó en revisión / con problema)
    estado_warning = None
    estado = creado.get("status")
    sub = creado.get("sub_status") or []
    if estado and estado != "active":
        motivos = ", ".join(sub) if sub else ""
        nombres = {"under_review": "en revisión por Mercado Libre",
                   "inactive": "inactiva (le falta algún dato obligatorio)",
                   "paused": "pausada"}
        estado_warning = (f"La publicación quedó {nombres.get(estado, estado)}"
                          + (f" — motivo: {motivos}" if motivos else "")
                          + ". Suele ser porque la categoría exige publicar por catálogo. "
                            "Si se da de baja, avisá para resolverlo.")
    warnings = [w for w in (foto_warning, desc_warning, estado_warning) if w]
    return jsonify({"ok": True, "id": item_id, "permalink": creado.get("permalink"),
                    "status": estado, "sub_status": sub,
                    "warning": " ".join(warnings) if warnings else None})


# ---------------------------------------------------------------- Tiendanube

@sync_bp.route("/tn/token", methods=["POST"])
def tn_token():
    """Canjea el code de OAuth de Tiendanube por un access token (no vence)."""
    body   = request.json or {}
    app_id = body.get("app_id", "").strip()
    secret = body.get("secret", "").strip()
    code   = body.get("code", "").strip()
    if not app_id or not secret or not code:
        return jsonify({"error": "Faltan App ID, Client Secret o code de Tiendanube"}), 400
    try:
        r = requests.post("https://www.tiendanube.com/apps/authorize/token",
                          data={"client_id": app_id, "client_secret": secret,
                                "grant_type": "authorization_code", "code": code},
                          headers={"Accept": "application/json"}, timeout=20)
    except requests.RequestException as e:
        return jsonify({"error": f"No se pudo contactar a Tiendanube: {e}"}), 502
    try:
        d = r.json()
    except ValueError:
        d = {}
    if r.status_code != 200 or not d.get("access_token"):
        detail = d.get("error_description") or d.get("error") or r.text[:300]
        return jsonify({"error": "Tiendanube rechazó la autorización", "detalle": detail}), 502
    return jsonify({"access_token": d["access_token"], "store_id": d.get("user_id")})


@sync_bp.route("/tn")
def get_tn():
    """Una página de productos de Tiendanube por llamada (la UI itera con
    progreso, para no exceder el timeout del servidor)."""
    store_id = request.args.get("store_id", os.environ.get("TN_STORE_ID", ""))
    token    = request.args.get("token", os.environ.get("TN_TOKEN", ""))
    page     = int(request.args.get("page", 1))
    if not store_id or not token:
        return jsonify({"error": "Faltan store_id o token de Tiendanube"}), 400
    try:
        r = requests.get(f"{TN_BASE}/{store_id}/products",
                         headers=tn_headers(token),
                         params={"page": page, "per_page": 200,
                                 "fields": "id,name,canonical_url,published,variants"},
                         timeout=30)
        if r.status_code == 404:  # última página
            return jsonify({"items": [], "has_more": False, "page": page})
        r.raise_for_status()
        products = r.json()
        items = []
        for p in products:
            resumen = {"id": p.get("id"), "name": tn_nombre(p.get("name")),
                       "published": p.get("published"), "url": p.get("canonical_url")}
            raices = sorted({sku_raiz(v.get("sku")) for v in p.get("variants") or []} - {None})
            items.append({"resumen": resumen, "raices": raices})
        return jsonify({"items": items, "has_more": len(products) == 200, "page": page})
    except requests.HTTPError as e:
        return jsonify({"error": f"Tiendanube: {e.response.status_code} {e.response.text[:300]}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def html_a_texto(html):
    """Convierte la descripción HTML de Tiendanube a texto plano, conservando
    saltos de línea y viñetas, para poder reusarla/editarla en el wizard."""
    if not html:
        return ""
    t = html
    t = re.sub(r"(?i)<\s*br\s*/?>", "\n", t)
    t = re.sub(r"(?i)</\s*(p|div|h[1-6])\s*>", "\n\n", t)
    t = re.sub(r"(?i)<\s*li[^>]*>", "- ", t)
    t = re.sub(r"(?i)</\s*li\s*>", "\n", t)
    t = re.sub(r"<[^>]+>", "", t)                    # resto de tags
    import html as _htmlmod
    t = _htmlmod.unescape(t)
    t = re.sub(r"[ \t]+\n", "\n", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


@sync_bp.route("/tn/fotos")
def tn_fotos():
    """Fotos y descripción de un producto ya publicado en Tiendanube (para
    reutilizarlas al publicar el mismo producto en ML)."""
    store_id   = request.args.get("store_id", os.environ.get("TN_STORE_ID", ""))
    token      = request.args.get("token", os.environ.get("TN_TOKEN", ""))
    product_id = request.args.get("product_id", "")
    if not store_id or not token or not product_id:
        return jsonify({"fotos": [], "error": "Faltan parámetros"})
    try:
        r = requests.get(f"{TN_BASE}/{store_id}/products/{product_id}",
                         headers=tn_headers(token),
                         params={"fields": "id,images,description"},
                         timeout=20)
        r.raise_for_status()
        p = r.json()
        desc = p.get("description") or {}
        # description viene por idioma {es: "<p>..."}; tomamos es o el primero
        desc_html = desc.get("es") if isinstance(desc, dict) else desc
        if isinstance(desc, dict) and not desc_html:
            desc_html = next((v for v in desc.values() if v), "")
        return jsonify({
            "fotos": [im.get("src") for im in p.get("images") or [] if im.get("src")],
            "descripcion": html_a_texto(desc_html),
        })
    except Exception as e:
        return jsonify({"fotos": [], "error": str(e)})


@sync_bp.route("/tn/reencuadrar", methods=["POST"])
def tn_reencuadrar():
    """Re-encuadra a 1740x1170 las fotos de un producto YA publicado en TN que
    no estén en esa medida (las que se subieron antes del arreglo y rompen la
    grilla). Reemplaza cada foto fuera de medida por su versión encuadrada."""
    body       = request.json or {}
    store_id   = body.get("store_id", os.environ.get("TN_STORE_ID", ""))
    token      = body.get("token", os.environ.get("TN_TOKEN", ""))
    product_id = body.get("product_id", "")
    if not store_id or not token or not product_id:
        return jsonify({"error": "Faltan store_id, token o product_id"}), 400
    try:
        r = requests.get(f"{TN_BASE}/{store_id}/products/{product_id}",
                         headers=tn_headers(token), params={"fields": "id,images"},
                         timeout=20)
        r.raise_for_status()
        imgs = sorted(r.json().get("images") or [], key=lambda x: x.get("position") or 0)
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el producto: {e}"}), 502

    arregladas, ya_ok, fallidas, ultimo_error = 0, 0, 0, None
    for im in imgs:
        src = im.get("src")
        if not src or not im.get("id"):
            continue
        try:
            contenido, ct, ext = descargar_imagen(src)
            # Aclarar un fondo negro (foto que quedó con transparencia aplanada
            # en negro) antes de decidir si hace falta reencuadrar
            aclarada = aclarar_fondo_negro(contenido)
            if aclarada:
                contenido = aclarada
            elif dim_imagen(contenido) == (1740, 1170):
                ya_ok += 1  # ya está en medida y sin fondo negro
                continue
            enc = encuadrar_1740x1170(contenido)
            if not enc:
                fallidas += 1
                continue
            nuevo, _, next_ext = enc
            # Agregar la versión encuadrada y borrar la vieja
            ra = requests.post(f"{TN_BASE}/{store_id}/products/{product_id}/images",
                               headers=tn_headers(token),
                               json={"attachment": base64.b64encode(nuevo).decode(),
                                     "filename": f"foto-{im.get('id')}.{next_ext}"},
                               timeout=60)
            if ra.status_code not in (200, 201):
                fallidas += 1
                ultimo_error = f"TN {ra.status_code}: {ra.text[:300]}"
                continue
            requests.delete(f"{TN_BASE}/{store_id}/products/{product_id}/images/{im['id']}",
                            headers=tn_headers(token), timeout=20)
            arregladas += 1
        except Exception as e:
            fallidas += 1
            ultimo_error = str(e)[:300]

    return jsonify({"ok": True, "arregladas": arregladas, "ya_ok": ya_ok,
                    "fallidas": fallidas, "detalle": ultimo_error})


@sync_bp.route("/tn/reemplazar-fotos", methods=["POST"])
def tn_reemplazar_fotos():
    """Reemplaza TODAS las fotos de un producto ya publicado en TN por las
    imágenes elegidas (encuadradas a 1740x1170). Sirve para arreglar un producto
    cuya foto quedó mal (ej. fondo negro) subiendo una nueva."""
    body       = request.json or {}
    store_id   = body.get("store_id", os.environ.get("TN_STORE_ID", ""))
    token      = body.get("token", os.environ.get("TN_TOKEN", ""))
    product_id = body.get("product_id", "")
    imagenes   = body.get("images") or []
    if not store_id or not token or not product_id:
        return jsonify({"error": "Faltan store_id, token o product_id"}), 400
    if not imagenes:
        return jsonify({"error": "No hay fotos elegidas para reemplazar"}), 400

    # Encuadrar las nuevas (fondo blanco), descartando chicas y repetidas
    nuevas, vistas = [], set()
    for im in imagenes:
        u = im.get("src")
        if not u:
            continue
        try:
            contenido, ct, ext = descargar_imagen(u)
            if not foto_calidad_ok(contenido):
                continue
            firma = firma_imagen(contenido)
            if firma in vistas:
                continue
            vistas.add(firma)
            enc = encuadrar_1740x1170(contenido, im.get("zoom", 0.92))
            if not enc:
                continue
            nuevo, _, next_ext = enc
            nuevas.append((base64.b64encode(nuevo).decode(), next_ext))
        except Exception:
            continue
    if not nuevas:
        return jsonify({"error": "Ninguna foto elegida sirvió (muy chicas o no se pudieron bajar)."}), 400

    # IDs de las fotos actuales (para borrarlas después de subir las nuevas)
    try:
        r = requests.get(f"{TN_BASE}/{store_id}/products/{product_id}",
                         headers=tn_headers(token), params={"fields": "id,images"}, timeout=20)
        r.raise_for_status()
        viejas = [im.get("id") for im in r.json().get("images") or [] if im.get("id")]
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el producto: {e}"}), 502

    subidas, ultimo_error = 0, None
    for i, (b64, ext) in enumerate(nuevas):
        try:
            ra = requests.post(f"{TN_BASE}/{store_id}/products/{product_id}/images",
                               headers=tn_headers(token),
                               json={"attachment": b64, "filename": f"foto-{i+1}.{ext}"},
                               timeout=60)
            if ra.status_code in (200, 201):
                subidas += 1
            else:
                ultimo_error = f"TN {ra.status_code}: {ra.text[:300]}"
        except Exception as e:
            ultimo_error = str(e)[:300]
    if not subidas:
        return jsonify({"error": "No se pudieron subir las fotos nuevas. " + (ultimo_error or "")}), 502
    # Recién ahora borramos las viejas (para no dejar el producto sin imágenes)
    for vid in viejas:
        try:
            requests.delete(f"{TN_BASE}/{store_id}/products/{product_id}/images/{vid}",
                            headers=tn_headers(token), timeout=20)
        except Exception:
            pass
    return jsonify({"ok": True, "subidas": subidas, "borradas": len(viejas)})


@sync_bp.route("/publish/tn", methods=["POST"])
def publish_tn():
    body     = request.json or {}
    store_id = body.get("store_id", os.environ.get("TN_STORE_ID", ""))
    token    = body.get("token", os.environ.get("TN_TOKEN", ""))
    product  = body.get("product")
    if not store_id or not token or not product:
        return jsonify({"error": "Faltan store_id, token o product"}), 400

    # Re-alojar fotos: se descargan acá y van a TN como adjunto base64, para
    # evitar "Remote image not found" en sitios que bloquean hotlink.
    # Se descartan las de baja calidad (se verían pixeladas) y las repetidas
    # (misma foto con URLs distintas), para no publicar la imagen 2, 3 o 4 veces.
    nuevas = []
    fotos_descartadas = 0
    vistas = {}      # firma -> posición (1-based) de la foto que quedó
    pos_por_src = {}  # src original -> posición en la lista final de imágenes
    for im in product.get("images") or []:
        u = im.get("src")
        if not u:
            continue
        try:
            contenido, ct, ext = descargar_imagen(u)
            if not foto_calidad_ok(contenido):
                fotos_descartadas += 1
                continue
            firma = firma_imagen(contenido)
            if firma in vistas:  # foto repetida: apunta a la que ya quedó
                pos_por_src[u] = vistas[firma]
                continue
            # Encuadrar a 1740x1170 (estándar Muvin en Tiendanube), respetando
            # el zoom elegido. Si el encuadre falla, NO se publica sin encuadrar.
            encuadrada = encuadrar_1740x1170(contenido, im.get("zoom", 0.92))
            if not encuadrada:
                fotos_descartadas += 1
                continue
            contenido, ct, ext = encuadrada
            nuevas.append({"attachment": base64.b64encode(contenido).decode(),
                           "filename": f"foto-{len(nuevas)+1}.{ext}"})
            vistas[firma] = len(nuevas)     # posición 1-based
            pos_por_src[u] = len(nuevas)
        except Exception:
            # No se pudo bajar/procesar: se descarta (no se publica sin encuadrar)
            fotos_descartadas += 1
    # Publicamos SOLO fotos ya encuadradas a 1740x1170 (nunca las originales).
    entraron = len(product.get("images") or [])
    foto_warning = None
    if nuevas:
        product["images"] = nuevas
        if fotos_descartadas:
            foto_warning = (f"Se descartaron {fotos_descartadas} foto(s) repetidas o "
                            "de baja calidad (muy chicas o borrosas).")
    else:
        # Ninguna foto se pudo procesar: no mandamos images:[] y avisamos fuerte
        product.pop("images", None)
        if entraron:
            foto_warning = ("⚠ Ninguna de las fotos elegidas se pudo publicar (todas "
                            "muy chicas: se necesitan ~1000px de lado). El producto se "
                            "publicó SIN fotos: subí una imagen más grande.")

    try:
        r = requests.post(f"{TN_BASE}/{store_id}/products",
                          headers=tn_headers(token), json=product, timeout=60)
    except requests.RequestException as e:
        return jsonify({"error": f"No se pudo contactar a Tiendanube: {e}"}), 502
    if r.status_code not in (200, 201):
        try:
            detail = r.json()
        except ValueError:
            detail = r.text[:500]
        return jsonify({"error": "Tiendanube rechazó la publicación", "detalle": detail}), 502
    creado = r.json()

    # Asignar la foto específica a cada variante (si el usuario cargó alguna).
    # TN asigna los ids de imagen al crear, así que recién ahora podemos
    # vincular cada variante con su imagen (por posición) vía un PUT.
    variant_fotos = body.get("variant_fotos") or {}
    if variant_fotos and creado.get("id"):
        try:
            id_por_pos = {im.get("position"): im.get("id")
                          for im in creado.get("images") or [] if im.get("id")}
            pid = creado["id"]
            for var in creado.get("variants") or []:
                src = variant_fotos.get(var.get("sku"))
                if not src:
                    continue
                img_id = id_por_pos.get(pos_por_src.get(src))
                if not img_id:
                    continue
                try:
                    requests.put(f"{TN_BASE}/{store_id}/products/{pid}/variants/{var['id']}",
                                 headers=tn_headers(token), json={"image_id": img_id},
                                 timeout=20)
                except requests.RequestException:
                    pass
        except Exception:
            pass  # el producto ya está creado; la asignación por variante es un extra

    return jsonify({"ok": True, "id": creado.get("id"),
                    "url": creado.get("canonical_url"),
                    "warning": foto_warning})


# ------------------------------------------------- borrador y descripciones

def titulo_desde_nombre(nombre):
    """'Candado, U-Lock - Kryptonite Evolution LS - Anaranjado de 12' ->
    'Candado U-Lock Kryptonite Evolution LS Anaranjado de 12'"""
    partes = [p.strip() for p in nombre.split(" - ") if p.strip()]
    if partes:
        partes[0] = partes[0].replace(",", "")
    titulo = " ".join(partes)
    titulo = re.sub(r"\s+", " ", titulo).strip()
    return titulo[:60].strip()


@sync_bp.route("/draft", methods=["POST"])
def draft():
    body   = request.json or {}
    nombre = body.get("nombre", "")
    alt    = (body.get("alt") or "").strip()  # código del fabricante (Hansa)
    titulo = titulo_desde_nombre(nombre)
    q = urllib.parse.quote_plus(((titulo or nombre) + (" " + alt if alt else "")).strip())
    return jsonify({
        "titulo": titulo,
        "photo_search": {
            "google": f"https://www.google.com/search?tbm=isch&q={q}",
            "bing": f"https://www.bing.com/images/search?q={q}",
            "duckduckgo": f"https://duckduckgo.com/?iax=images&ia=images&q={q}",
        },
    })


ESTILO_MUVIN = """\
¿Desplazamiento diario entre semana o aventuras de fin de semana? ¿Una aventura \
de tres días o una vuelta al mundo de 12 meses? La Four Corners 1 ahora viene con \
una nueva transmisión 2x9 MicroShift Sword que ofrece mayor fiabilidad y una gama \
más amplia de marchas. El cuadro y la horquilla de acero CroMo 4130 conificados \
están diseñados para ser cómodos en terrenos difíciles, pero también para viajes \
con carga completa. Hemos incluido seis soportes para botellas, ojales para \
portabultos y guardabarros, soportes para horquilla lowrider, amplio espacio libre \
para neumáticos y frenos de disco para que puedas afrontar cualquier condición y \
terreno.

La vida se trata del viaje, no del destino, y para eso está la Four Corners.

- Cuadro: CrMo Serie 1, geometría biométrica, soportes para guardabarros y portapaquetes
- Horquilla: Serie 1 CrMo, ojales para portabotellas, montaje de disco IS
- Frenos: Disco mecánico de carretera Tektro Spyre-C, rotor de 160 mm
- Cadena: KMC X9"""

PROMPT_SISTEMA = """Sos el redactor de Muvin (muvin.com.ar), una tienda argentina \
de bicicletas urbanas y plegables, movilidad eléctrica y accesorios de ciclismo.

Escribís descripciones de producto CONCISAS en español rioplatense, con esta estructura:
1. UN párrafo corto (2 a 4 oraciones) que conecta el producto con su uso real. \
Cálido pero directo, sin emojis, sin superlativos vacíos, sin relleno.
2. Una lista de especificaciones clave, cada línea "- Componente: detalle". \
Solo los datos que importan; no infles la lista.

Priorizá los datos técnicos concretos por sobre la prosa: es mejor una \
descripción corta con datos reales que un texto largo y vago.

Ejemplo del TONO (no de la extensión — para un accesorio simple, mucho más corto):
---
{ejemplo}
---

Tenés acceso a búsqueda web. ANTES de escribir, investigá el producto real:
- Buscá por marca + modelo + código de fabricante (MPN) en el sitio oficial \
de la marca; si no, en tiendas o reseñas serias del rubro ciclismo.
- Si te doy la URL de la página del producto, consultala.
- Basá las especificaciones en la ficha técnica real que encuentres \
(materiales, medidas, peso, compatibilidad, características, usos).

Reglas estrictas:
- Usá SOLO datos que puedas verificar en las fuentes o que te haya dado el \
usuario. Si no encontrás un dato, omitilo — NUNCA inventes specs.
- Si no encontrás nada del producto, escribí una descripción general honesta \
con lo que sí sabés, sin rellenar con datos inventados.
- No menciones precio, stock ni envío. No cites las URLs en el texto final.
- NO incluyas el código de fabricante / MPN en la descripción (ni en la \
lista de especificaciones ni en el texto): un sistema lo agrega solo al final.

IMPORTANTE — FORMATO DE SALIDA: podés pensar o buscar lo que necesites, pero \
la descripción final va OBLIGATORIAMENTE encerrada entre las etiquetas \
<desc> y </desc>, sin nada más adentro (ni comentarios, ni el código, ni \
"aquí está"). Ejemplo:
<desc>
[primer párrafo]

- Componente: detalle
- Componente: detalle
</desc>"""


def descripcion_plantilla(nombre, titulo):
    partes = [p.strip() for p in (nombre or "").split(" - ") if p.strip()]
    categoria = re.sub(r"\s+", " ", partes[0].replace(",", " ")).strip() if partes else "Producto"
    detalle = partes[2] if len(partes) > 2 else ""
    marca_modelo = partes[1] if len(partes) > 1 else titulo
    lineas = [
        f"{titulo}.",
        "",
        f"Sumá a tu bici un {categoria.lower()} pensado para el uso urbano de todos los días.",
        "",
        f"- Producto: {categoria}",
        f"- Marca y modelo: {marca_modelo}",
    ]
    if detalle:
        lineas.append(f"- Detalle: {detalle}")
    return "\n".join(lineas)


def texto_a_html(texto):
    html_partes = []
    bullets = []
    for linea in texto.splitlines():
        s = linea.strip()
        if s.startswith("- "):
            bullets.append(f"<li>{s[2:].strip()}</li>")
            continue
        if bullets:
            html_partes.append("<ul>" + "".join(bullets) + "</ul>")
            bullets = []
        if s:
            html_partes.append(f"<p>{s}</p>")
    if bullets:
        html_partes.append("<ul>" + "".join(bullets) + "</ul>")
    return "".join(html_partes)


_NARRACION = re.compile(
    r"^\s*(voy a |déjame|dejame|let me|i'?ll |i will |now |the result|the mpn|"
    r"the product|based on|según la ficha|tengo la ficha|escribo la|busqué|"
    r"encontré|investigar|primero,|first,|el mpn|el código .* corresponde|"
    r"perfecto[.,]|listo[.,]|aquí|acá va|here'?s|this ).*", re.I)


_CIERRE_NARRACION = re.compile(
    r"(redacto|escribo|redacción|aqu[ií] (va|est[áa])|acá va|here'?s)", re.I)


def limpiar_narracion(texto):
    """Saca líneas de 'pensamiento' que la IA pueda dejar al principio."""
    lineas = texto.split("\n")
    # Si en las primeras líneas hay un cierre tipo "Redacto la descripción",
    # descartar todo hasta ahí (es todo narración previa)
    for i, l in enumerate(lineas[:6]):
        if _CIERRE_NARRACION.search(l) and l.strip().endswith((".", ":")):
            lineas = lineas[i + 1:]
            break
    while lineas and (not lineas[0].strip() or _NARRACION.match(lineas[0])):
        lineas.pop(0)
    return "\n".join(lineas).strip()


_LINEA_CODIGO = re.compile(
    r"^\s*[-•*]?\s*(c[óo]digo(\s+de\s+fabricante|\s+alternativo)?|mpn|"
    r"n[úu]mero\s+de\s+pieza|part\s*number)\s*[:\-].*$", re.I)


def con_codigo(texto, codigo):
    """Deja una única línea final 'Código: <alternativo>', quitando cualquier
    mención del código que la IA haya puesto en el texto. Si no hay código
    alternativo, no agrega nada (pero igual limpia las menciones sueltas)."""
    lineas = [l for l in texto.splitlines() if not _LINEA_CODIGO.match(l)]
    texto = "\n".join(lineas).rstrip()
    if codigo:
        texto += f"\n\nCódigo: {codigo}"
    return texto


@sync_bp.route("/describe", methods=["POST"])
def describe():
    body    = request.json or {}
    nombre  = body.get("nombre", "")
    titulo  = body.get("titulo") or titulo_desde_nombre(nombre)
    datos    = body.get("datos", "")  # specs/notas extra que cargue el usuario
    alt      = (body.get("alt") or "").strip()   # código del fabricante
    marca    = (body.get("marca") or "").strip()
    slug_url = (body.get("slug_url") or "").strip()  # página del fabricante
    sku      = (body.get("codigo") or "").strip()  # vacío => sin línea Código
    api_key  = body.get("api_key") or os.environ.get("ANTHROPIC_API_KEY", "")

    if not api_key:
        texto = con_codigo(descripcion_plantilla(nombre, titulo), sku)
        return jsonify({"texto": texto, "html": texto_a_html(texto),
                        "generado_con_ia": False,
                        "warning": "Sin API key de Anthropic: se usó una plantilla básica. "
                                   "Cargá una key en Configuración para generar con IA."})
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        contenido = f"Nombre interno del producto (ERP): {nombre}\nTítulo de la publicación: {titulo}"
        if marca:
            contenido += f"\nMarca: {marca}"
        if alt:
            contenido += f"\nCódigo del fabricante (MPN) para identificar el modelo exacto: {alt}"
        if slug_url:
            contenido += f"\nPágina oficial del producto (consultala): {slug_url}"
        if datos:
            contenido += f"\nDatos adicionales del usuario:\n{datos}"

        # Equilibrado: Sonnet + menos investigación (más barato y rápido que
        # Opus, calidad muy buena). Configurable por variable de entorno.
        tools = [{"type": "web_search_20260209", "name": "web_search", "max_uses": 2},
                 {"type": "web_fetch_20260209", "name": "web_fetch", "max_uses": 2}]
        model = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-6")

        def generar(tools_param):
            mensajes = [{"role": "user", "content": contenido}]
            for _ in range(6):  # continúa si el loop de web search se pausa
                resp = client.messages.create(
                    model=model, max_tokens=3000,
                    system=PROMPT_SISTEMA.format(ejemplo=ESTILO_MUVIN),
                    messages=mensajes, tools=tools_param)
                if resp.stop_reason == "pause_turn":
                    mensajes.append({"role": "assistant", "content": resp.content})
                    continue
                return resp
            return resp

        try:
            resp = generar(tools)
        except Exception:
            resp = generar([])  # si la búsqueda web no está disponible, sin tools

        # La descripción real es el texto que viene DESPUÉS de la última
        # búsqueda; lo anterior es la IA narrando su proceso ("voy a buscar…")
        bloques = list(resp.content)
        ult_tool = -1
        for i, b in enumerate(bloques):
            if getattr(b, "type", "") not in ("text", "thinking"):
                ult_tool = i
        finales = [b for b in bloques[ult_tool + 1:] if getattr(b, "type", "") == "text"]
        texto = "\n".join(b.text for b in finales).strip()
        if not texto:  # sin herramientas usadas: tomar todo el texto
            texto = "\n".join(b.text for b in bloques
                              if getattr(b, "type", "") == "text").strip()
        if not texto:
            raise RuntimeError("La API no devolvió texto")
        # La descripción real va entre <desc>...</desc>; si está, tomamos solo
        # eso (evita cualquier narración fuera de las etiquetas)
        m = re.search(r"<desc>\s*(.*?)\s*</desc>", texto, re.S | re.I)
        texto = m.group(1) if m else limpiar_narracion(texto)
        texto = con_codigo(texto, sku)
        return jsonify({"texto": texto, "html": texto_a_html(texto), "generado_con_ia": True})
    except Exception as e:
        texto = con_codigo(descripcion_plantilla(nombre, titulo), sku)
        return jsonify({"texto": texto, "html": texto_a_html(texto),
                        "generado_con_ia": False,
                        "warning": f"Falló la generación con IA ({e}); se usó la plantilla."})
